"""Excel (.xlsx) document parser using openpyxl."""

import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse_xlsx(file_path: str, source: str = "") -> dict:
    """Convert an Excel workbook to a readable text representation.

    Each sheet is rendered with its name as a heading, followed by rows
    with column headers as labels.
    Returns: {"text": str, "metadata": dict}
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sections = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        sheet_lines = [f"## {sheet_name}"]

        for row in rows[1:]:
            parts = []
            for h, v in zip(headers, row):
                if v is not None and str(v).strip():
                    parts.append(f"{h}: {v}")
            if parts:
                sheet_lines.append(" | ".join(parts))
                total_rows += 1

        sections.append("\n".join(sheet_lines))

    wb.close()
    full_text = "\n\n".join(sections)

    return {
        "text": full_text,
        "metadata": {
            "source": source,
            "format": "xlsx",
            "sheet_count": len(wb.sheetnames),
            "row_count": total_rows,
        },
    }
